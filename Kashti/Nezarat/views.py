from django.shortcuts import render, redirect, get_object_or_404
from django.forms import modelformset_factory
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.forms import AuthenticationForm
from .models import (Company, TransportItem,  Employee, JobApplication, My_Employee, Part,
                     CustomAttribute, Ship, ShipPart, Message, User, Warehouse,DamageReport , WarehouseSubStorage ,
                     Reminder , TransportOperation, ShipSubWarehouse)
from django.contrib import messages
from .forms import (CustomUserCreationForm, CompanyCreationForm, Choice, PartForm, CustomAttributeForm, 
                    ShipForm, ShipPartForm, TransportItemForm ,
                    SendMessage, WarehouseForm, TransportOperationForm ,
                  ReminderForm, WarehouseSubStorageForm,
                 ShipSubWarehouseForm , ShipAttributeForm , DamageReportForm, WarehouseSubStorageFormSet
                 )
import logging
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse , HttpResponseForbidden
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.db import transaction
from openpyxl.styles import PatternFill
from django.utils import timezone
import csv
from django.http import HttpResponse
from xhtml2pdf import pisa
from django.contrib import messages as django_messages  # برای نوتیف
from .models import Message  # مدل چت
from django.contrib.auth import get_user_model
from django.contrib import messages as django_messages
from django.http import HttpResponseForbidden, JsonResponse

User = get_user_model()


logger = logging.getLogger(__name__)

def base_context(request):
    if request.user.is_authenticated and hasattr(request.user, 'company'):
        now = timezone.now()
        reminders = Reminder.objects.filter(company_code=request.user.company.company_code, reminder_date__lte=now)
        return {'notifications': reminders}
    return {}

def get_company(user):
    try:
        return Company.objects.get(user=user)
    except Company.DoesNotExist:
        return None

@login_required
def home(request):
    try:
        company = Company.objects.get(user=request.user)
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید. لطفاً یک شرکت ثبت کنید.')
        return redirect('sign_up')

    ships = Ship.objects.filter(company_code=company.company_code)
    current_ship = ships.first()
    current_time = timezone.now()
    expired_parts = Part.objects.filter(expiry_date__lt=current_time, company_code=company.company_code)
    searched_ship = None
    if request.method == 'POST':
        query = request.POST.get('search_query')
        searched_ship = ships.filter(name__icontains=query).first()
    warehouses = Warehouse.objects.filter(company_code=company.company_code)
    context = {
        'company': company,
        'current_ship': current_ship,
        'searched_ship': searched_ship,
        'ships': ships,
        'shipparts': ShipPart.objects.filter(ship__company_code=company.company_code),
        'customs': CustomAttribute.objects.filter(company_code=company.company_code),
        'expired_parts': expired_parts,
        'warehouses': warehouses,
        'message': '',
    }
    return render(request, 'home.html', context)

def register_company(request):
    if request.method == 'POST':
        form = CompanyCreationForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password1'])
            user.title = 'C'
            user.save()
            Company.objects.create(user=user, name=request.POST['name'], company_code=request.POST['company_code'])
            messages.success(request, 'شرکت شما ثبت شد.')
            login(request, user)
            return redirect('home')
    else:
        form = CompanyCreationForm()
    return render(request, 'register_company.html', {'form': form})

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('home')
    else:
        form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})

def logout_user(request):
    logout(request)
    messages.success(request, "با موفقیت خارج شدید")
    return redirect("home")

def register_employee(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST, request.FILES)
        if form.is_valid():
            company_code = form.cleaned_data['company_code']
            try:
                company = Company.objects.get(company_code=company_code)
                user = form.save(commit=False)
                user.set_password(form.cleaned_data['password1'])
                user.title = 'E'
                user.save()
                login(request, user)
                picture = form.cleaned_data.get('profile_picture')
                Employee.objects.create(user=user, company_code=company_code, profile_picture=picture)
                JobApplication.objects.create(employee=Employee.objects.get(user=user), company=company)
                messages.success(request, 'درخواست استخدام شما با موفقیت ثبت شد.')
                return redirect('home')
            except Company.DoesNotExist:
                messages.error(request, 'کد شرکت نامعتبر است.')
                return redirect('register_employee')
    else:
        form = CustomUserCreationForm()
    return render(request, 'signupE.html', {'form': form})

@login_required
def ship_list(request):
    try:
        company = Company.objects.get(user=request.user)
        ships = Ship.objects.filter(company_code=company.company_code)
        return render(request, 'ship_list.html', {'ships': ships})
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')

# ویوهای به‌روزرسانی و حذف کشتی
@login_required
def update_ship(request, ship_id):
    try:
        company = Company.objects.get(user=request.user)
        ship = get_object_or_404(Ship, id=ship_id, company_code=company.company_code)
        if request.method == 'POST':
            form = ShipForm(request.POST, request.FILES, instance=ship)
            if form.is_valid():
                ship = form.save(commit=False)
                ship.company_code = company.company_code
                ship.save()
                messages.success(request, 'کشتی به‌روزرسانی شد.')
                return redirect('ship_list')
        else:
            form = ShipForm(instance=ship)
        return render(request, 'update_ship.html', {'form': form, 'ship': ship})
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')

@login_required
def delete_ship(request, ship_id):
    try:
        company = Company.objects.get(user=request.user)
        ship = get_object_or_404(Ship, id=ship_id, company_code=company.company_code)
        if request.method == 'POST':
            ship.delete()
            messages.success(request, 'کشتی حذف شد.')
            return redirect('ship_list')
        return render(request, 'delete_ship.html', {'ship': ship})
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')

def company_dashboard(request):
    try:
        company = Company.objects.get(user=request.user)
        applications = JobApplication.objects.filter(company=company)
        if request.method == 'POST':
            application_id = request.POST.get('application_id')
            action = request.POST.get('action')
            position = request.POST.get('position')
            application = JobApplication.objects.get(id=application_id)
            employee = Employee.objects.get(user=application.employee.user)

            # ست کردن position
            employee.user.position = position
            employee.user.save()

            if action == 'accept':
                # اگر فیلدی بنام is_accepted ندارید از .delete() یا status استفاده کنید.
                application.status = 'Accepted'
                application.save()
                My_Employee.objects.create(employee=employee, company=company, position=position)
                application.delete()
                return redirect('company_dashboard')

            elif action == 'reject':
                application.delete()
                employee.user.delete()  # یا فقط یوزر را حذف نکنید، بسته به نیاز
                return redirect('company_dashboard')

        return render(request, 'dashboard.html', {'applications': applications})
    except Company.DoesNotExist:
        return render(request, 'dashboard.html', {'message': 'شما هنوز شرکتی ثبت نکرده‌اید.'})

def sign_up(request):
    return render(request, 'sign_up.html')

@login_required
def myemployee(request):
    try:
        company = Company.objects.get(user=request.user)
        employees = My_Employee.objects.filter(company=company)
        return render(request, 'my_employee.html', {'employees': employees, 'company': company})
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')

@login_required
def mycolleagues(request):
    try:
        company = Company.objects.get(user=request.user)
        employees = Employee.objects.filter(company_code=company.company_code)
        return render(request, 'my_employee.html', {'employees': employees, 'company': company})
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')

@login_required
def nezarat(request):
    """
    صفحهٔ نظارت جامع:
    - تمام اطلاعات مورد نیاز پروژه را در یک صفحه تب‌بندی نمایش می‌دهد.
    - با امکان نمایش جداول، جستجو، ویرایش مودالی، نمودار آماری و غیره.
    """
    try:
        # پیدا کردن شرکت کاربر فعلی
        company = Company.objects.get(user=request.user)
        company_code = company.company_code

        # 1) لیست کارمندان (My_Employee) مربوط به همین شرکت
        employees = My_Employee.objects.filter(company=company)

        # 2) کشتی‌ها
        ships = Ship.objects.filter(company_code=company_code)

        # 3) قطعات نصب‌شده روی کشتی (ShipPart)
        ship_parts = ShipPart.objects.filter(ship__company_code=company_code)

        # 4) قطعات کلی (Part)
        parts = Part.objects.filter(company_code=company_code)

        # 5) انبارهای شرکت (Warehouse)
        warehouses = Warehouse.objects.filter(company_code=company_code)

        # 6) زیرانبارهای انبار (WarehouseSubStorage)
        subwarehouses = WarehouseSubStorage.objects.filter(warehouse__company_code=company_code)

        # 7) زیرانبارهای کشتی (ShipSubWarehouse)
        ship_subwarehouses = ShipSubWarehouse.objects.filter(ship__company_code=company_code)

        # 8) یادآوری‌ها
        reminders = Reminder.objects.filter(company_code=company_code)

        # 9) سفرهای شرکت (TransportOperation)
        transport_ops = TransportOperation.objects.filter(company_code=company_code)

        # پیدا کردن قطعات تاریخ انقضاشده
        expired_parts = parts.filter(expiry_date__lt=timezone.now())
        expired_message = None
        if expired_parts.exists():
            names = ", ".join([p.name for p in expired_parts])
            expired_message = f"⏰ عمر مفید قطعات زیر به پایان رسیده است: {names}"

        # محاسبه آمار برای نمودار
        ship_count = ships.count()
        warehouse_count = warehouses.count()
        part_count = parts.count()
        transport_count = transport_ops.count()
        employee_count = employees.count()

        stats = {
            'ships': ship_count,
            'warehouses': warehouse_count,
            'parts': part_count,
            'transports': transport_count,
            'employees': employee_count,
        }

        context = {
            'company': company,
            'employees': employees,
            'ships': ships,
            'ship_parts': ship_parts,
            'parts': parts,
            'warehouses': warehouses,
            'subwarehouses': subwarehouses,
            'ship_subwarehouses': ship_subwarehouses,
            'reminders': reminders,
            'transport_ops': transport_ops,
            'expired_message': expired_message,
            'stats': stats,
        }
        return render(request, 'nezarat.html', context)

    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')



@login_required
def nezaratc(request):
    try:
        company = Company.objects.get(user=request.user)
        company_code = company.company_code
        ship_queryset = Ship.objects.filter(company_code=company_code)
        part_queryset = Part.objects.filter(company_code=company_code)
        form = ShipPartForm()
        if 'add_ship_part' in request.POST:
            form = ShipPartForm(request.POST)
            if form.is_valid():
                ship_part = form.save(commit=False)
                part = ship_part.part
                if part.quantity >= ship_part.quantity:
                    part.quantity -= ship_part.quantity
                    part.save()
                    ship_part.save()
                    return redirect('nezaratc')
                else:
                    return render(request, 'nezaratc.html', {'form': form})
        form.fields['ship'].queryset = ship_queryset
        form.fields['part'].queryset = part_queryset
        return render(request, 'nezaratc.html', {'form': form})
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')

@login_required
def chat_view(request, receiver_id):
    try:
        company = Company.objects.get(user=request.user)
        receiver = get_object_or_404(User, id=receiver_id)



        # پیام‌های خوانده نشده رو بخونیم

        # ارسال پیام
        if request.method == 'POST':
            content = request.POST.get('content')
            if content:
                Message.objects.create(sender=request.user, receiver=receiver, content=content)
                return redirect('chat', receiver_id=receiver_id)

        # همه پیام‌ها
        messages_qs = Message.objects.filter(sender=request.user, receiver=receiver).union(
            Message.objects.filter(sender=receiver, receiver=request.user)).order_by('timestamp')

        return render(request, 'chat.html', {
            'receiver': receiver,
            'messages': messages_qs
        })

    except Company.DoesNotExist:
        django_messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')


@login_required
def export_searched_ship_to_excel(request):
    try:
        company = Company.objects.get(user=request.user)
        ship_id = request.GET.get('ship_id')
        ship = get_object_or_404(Ship, id=ship_id, company_code=company.company_code)
        parts = Part.objects.filter(company_code=company.company_code, warehouse__company_code=company.company_code)

        wb = Workbook()
        ws = wb.active
        ws.title = "اطلاعات کشتی"

        # استایل‌ها
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        # عنوان‌ها
        ws.append(["نام کشتی", "کد شرکت", "قطعات", "نام قطعه", "تعداد", "انبار"])

        for cell in ws["1:1"]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # اطلاعات کشتی
        ws.append([ship.name, ship.company_code, "", "", "", ""])

        # اضافه کردن اطلاعات قطعات
        for part in parts:
            if part.company_code == ship.company_code:
                ws.append([None, None, part.name, part.quantity, part.warehouse.name])
        
        # تنظیم اندازه ستون‌ها
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 25

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{ship.name}_info.xlsx"'
        wb.save(response)
        return response

    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')

@login_required
def update_shippart_quantity(request, shippart_id):
    try:
        company = Company.objects.get(user=request.user)
        shippart = get_object_or_404(ShipPart, id=shippart_id, ship__company_code=company.company_code)
        if request.method == 'POST':
            new_quantity = int(request.POST.get('quantity'))
            if new_quantity < 0:
                messages.error(request, 'تعداد نمی‌تواند منفی باشد.')
                return redirect('home')
            difference = shippart.quantity - new_quantity
            shippart.quantity = new_quantity
            shippart.save()
            if difference > 0:
                part = shippart.part
                part.quantity += difference
                part.save()
            messages.success(request, 'موجودی قطعه در کشتی به‌روزرسانی شد.')
            return redirect('home')
        return render(request, 'update_shippart_quantity.html', {'shippart': shippart})
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')

@login_required
def set_reminder(request, part_id):
    try:
        company = Company.objects.get(user=request.user)
        part = get_object_or_404(Part, id=part_id, company_code=company.company_code)
        if request.method == 'POST':
            form = ReminderForm(request.POST)
            if form.is_valid():
                reminder = form.save(commit=False)
                reminder.part = part
                reminder.company_code = company.company_code
                reminder.save()
                messages.success(request, 'یادآوری تنظیم شد.')
                return redirect('warehouse_details', warehouse_id=part.warehouse.id)
        else:
            form = ReminderForm()
        return render(request, 'set_reminder.html', {'form': form, 'part': part})
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')

# ویوهای به‌روزرسانی و حذف انبار
@login_required
def create_warehouse(request):
    try:
        company = Company.objects.get(user=request.user)
        if request.method == 'POST':
            form = WarehouseForm(request.POST)
            if form.is_valid():
                warehouse = form.save(commit=False)
                warehouse.company_code = company.company_code
                warehouse.save()
                messages.success(request, 'انبار با موفقیت ایجاد شد.')
                return redirect('warehouse_list')
        else:
            form = WarehouseForm()
        return render(request, 'create_warehouse.html', {'form': form})
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')

@login_required
def update_warehouse(request, warehouse_id):
    try:
        company = Company.objects.get(user=request.user)
        warehouse = get_object_or_404(Warehouse, id=warehouse_id, company_code=company.company_code)
        if request.method == 'POST':
            form = WarehouseForm(request.POST, instance=warehouse)
            if form.is_valid():
                warehouse = form.save(commit=False)
                warehouse.company_code = company.company_code
                warehouse.save()
                messages.success(request, 'انبار به‌روزرسانی شد.')
                return redirect('warehouse_list')
        else:
            form = WarehouseForm(instance=warehouse)
        return render(request, 'update_warehouse.html', {'form': form, 'warehouse': warehouse})
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')

@login_required
def delete_warehouse(request, warehouse_id):
    try:
        company = Company.objects.get(user=request.user)
        warehouse = get_object_or_404(Warehouse, id=warehouse_id, company_code=company.company_code)
        if request.method == 'POST':
            warehouse.delete()
            messages.success(request, 'انبار حذف شد.')
            return redirect('warehouse_list')
        return render(request, 'delete_warehouse.html', {'warehouse': warehouse})
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')

# ویوهای به‌روزرسانی و حذف کالا (Part)

def update_part(request, id):
    part = get_object_or_404(Part, id=id)
    if request.method == 'POST':
        form = PartForm(request.POST, instance=part)
        if form.is_valid():
            form.save()
            return redirect('nezarat')  # برگشت به صفحه اصلی پس از ذخیره
    else:
        form = PartForm(instance=part)
    return render(request, 'update_part.html', {'form': form, 'part': part})

@login_required
def delete_part(request, part_id):
    try:
        company = Company.objects.get(user=request.user)
        part = get_object_or_404(Part, id=part_id, company_code=company.company_code)
        if request.method == 'POST':
            warehouse_id = part.warehouse.id
            part.delete()
            messages.success(request, 'قطعه حذف شد.')
            return redirect('warehouse_details', warehouse_id=warehouse_id)
        return render(request, 'delete_part.html', {'part': part})
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')

@login_required
def warehouse_list(request):
    user = request.user
    try:
        company = Company.objects.get(user=user)
        warehouses = Warehouse.objects.filter(company_code=company.company_code)
    except Company.DoesNotExist:
        warehouses = Warehouse.objects.none()
    
    return render(request, 'warehouse_list.html', {'warehouses': warehouses})

def warehouse_details(request, warehouse_id):
    """جزئیات یک انبار + زیرانبارهای وابسته + نمایش قطعات هر زیرانبار"""
    warehouse = get_object_or_404(Warehouse, id=warehouse_id)

    # چک کنید این انبار متعلق به شرکت فعلی باشد
    try:
        company = Company.objects.get(user=request.user)
        if warehouse.company_code != company.company_code:
            return HttpResponseForbidden("شما اجازه دسترسی به این انبار را ندارید.")
    except Company.DoesNotExist:
        return HttpResponseForbidden("شما شرکت ثبت نکرده‌اید.")

    # زیرانبارهای مربوط به این انبار
    substorages = WarehouseSubStorage.objects.filter(warehouse=warehouse)
    # قطعات هر زیرانبار را جداگانه برای راحتی نمایش قالب جمع‌آوری می‌کنیم
    storage_parts = {
        sub: Part.objects.filter(warehouse_sub_storage=sub)
        for sub in substorages
    }

    context = {
        'warehouse': warehouse,
        'substorages': substorages,
        'storage_parts': storage_parts,
    }
    return render(request, 'warehouse_details.html', context)


# ویو لیست سفرها برای رهگیری

@login_required
def add_part_to_warehouse(request):
    try:
        company = Company.objects.get(user=request.user)
        if request.method == 'POST':
            form = PartForm(request.POST, request.FILES)
            if form.is_valid():
                part = form.save(commit=False)
                part.company_code = company.company_code
                warehouse_id = part.warehouse.id
                part.save()
                return redirect('warehouse_details', warehouse_id=warehouse_id)
        else:
            form = PartForm()
        return render(request, 'add_part_to_warehouse.html', {'form': form})
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')

@login_required
def update_ship(request, ship_id):
    try:
        company = Company.objects.get(user=request.user)
        ship = get_object_or_404(Ship, id=ship_id, company_code=company.company_code)
        if request.method == 'POST':
            form = ShipForm(request.POST, request.FILES, instance=ship)
            if form.is_valid():
                ship = form.save(commit=False)
                ship.company_code = company.company_code
                ship.save()
                messages.success(request, 'کشتی به‌روزرسانی شد.')
                return redirect('home')
        else:
            form = ShipForm(instance=ship)
        return render(request, 'update_ship.html', {'form': form, 'ship': ship})
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')


@login_required
def delete_ship(request, ship_id):
    try:
        company = Company.objects.get(user=request.user)
        ship = get_object_or_404(Ship, id=ship_id, company_code=company.company_code)
        if request.method == 'POST':
            ship.delete()
            messages.success(request, 'کشتی حذف شد.')
            return redirect('home')
        return render(request, 'delete_ship.html', {'ship': ship})
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')


@login_required
def update_warehouse(request, warehouse_id):
    try:
        company = Company.objects.get(user=request.user)
        warehouse = get_object_or_404(Warehouse, id=warehouse_id, company_code=company.company_code)
        if request.method == 'POST':
            form = WarehouseForm(request.POST, instance=warehouse)
            if form.is_valid():
                warehouse = form.save(commit=False)
                warehouse.company_code = company.company_code
                warehouse.save()
                messages.success(request, 'انبار به‌روزرسانی شد.')
                return redirect('home')
        else:
            form = WarehouseForm(instance=warehouse)
        return render(request, 'update_warehouse.html', {'form': form, 'warehouse': warehouse})
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')


@login_required
def delete_warehouse(request, warehouse_id):
    try:
        company = Company.objects.get(user=request.user)
        warehouse = get_object_or_404(Warehouse, id=warehouse_id, company_code=company.company_code)
        if request.method == 'POST':
            warehouse.delete()
            messages.success(request, 'انبار حذف شد.')
            return redirect('home')
        return render(request, 'delete_warehouse.html', {'warehouse': warehouse})
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')
    
from django.shortcuts import get_object_or_404


@login_required
def update_reminder(request, reminder_id):
    try:
        company = Company.objects.get(user=request.user)
        reminder = get_object_or_404(Reminder, id=reminder_id, company_code=company.company_code)
        if request.method == 'POST':
            form = ReminderForm(request.POST, instance=reminder, user=request.user)
            if form.is_valid():
                form.save()
                messages.success(request, 'یادآوری به‌روزرسانی شد.')
                return redirect('reminder_list')
        else:
            form = ReminderForm(instance=reminder, user=request.user)
        return render(request, 'update_reminder.html', {'form': form})
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')

@login_required
def delete_reminder(request, reminder_id):
    try:
        company = Company.objects.get(user=request.user)
        reminder = get_object_or_404(Reminder, id=reminder_id, company_code=company.company_code)
        if request.method == 'POST':
            reminder.delete()
            messages.success(request, 'یادآوری حذف شد.')
            return redirect('reminder_list')
        return render(request, 'delete_reminder.html', {'reminder': reminder})
    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')

# -----------------------------
# 🚢 مدیریت کشتی‌ها
# -----------------------------

@login_required
def create_ship(request):
    company = get_object_or_404(Company, user=request.user)
    if request.method == 'POST':
        form = ShipForm(request.POST, request.FILES)
        if form.is_valid():
            ship = form.save(commit=False)
            ship.company_code = company.company_code
            ship.save()
            messages.success(request, 'کشتی جدید ایجاد شد.')
            return redirect('home')
    else:
        form = ShipForm()
    return render(request, 'create_ship.html', {'form': form})

@login_required
def update_ship(request, ship_id):
    ship = get_object_or_404(Ship, id=ship_id, company_code=request.user.company.company_code)
    if request.method == 'POST':
        form = ShipForm(request.POST, request.FILES, instance=ship)
        if form.is_valid():
            form.save()
            messages.success(request, 'اطلاعات کشتی به‌روزرسانی شد.')
            return redirect('home')
    else:
        form = ShipForm(instance=ship)
    return render(request, 'update_ship.html', {'form': form, 'ship': ship})

@login_required
def delete_ship(request, ship_id):
    ship = get_object_or_404(Ship, id=ship_id, company_code=request.user.company.company_code)
    if request.method == 'POST':
        ship.delete()
        messages.success(request, 'کشتی حذف شد.')
        return redirect('home')
    return render(request, 'delete_ship.html', {'ship': ship})

# -----------------------------
# 📦 مدیریت انبارها
# -----------------------------

@login_required
def create_warehouse(request):
    company = get_object_or_404(Company, user=request.user)
    if request.method == 'POST':
        form = WarehouseForm(request.POST)
        if form.is_valid():
            warehouse = form.save(commit=False)
            warehouse.company_code = company.company_code
            warehouse.save()
            messages.success(request, 'انبار جدید ایجاد شد.')
            return redirect('warehouse_list')
    else:
        form = WarehouseForm()
    return render(request, 'create_warehouse.html', {'form': form})

@login_required
def update_warehouse(request, warehouse_id):
    warehouse = get_object_or_404(Warehouse, id=warehouse_id, company_code=request.user.company.company_code)
    if request.method == 'POST':
        form = WarehouseForm(request.POST, instance=warehouse)
        if form.is_valid():
            form.save()
            messages.success(request, 'اطلاعات انبار به‌روزرسانی شد.')
            return redirect('warehouse_list')
    else:
        form = WarehouseForm(instance=warehouse)
    return render(request, 'update_warehouse.html', {'form': form, 'warehouse': warehouse})

@login_required
def delete_warehouse(request, warehouse_id):
    warehouse = get_object_or_404(Warehouse, id=warehouse_id, company_code=request.user.company.company_code)
    if request.method == 'POST':
        warehouse.delete()
        messages.success(request, 'انبار حذف شد.')
        return redirect('warehouse_list')
    return render(request, 'delete_warehouse.html', {'warehouse': warehouse})

# -----------------------------
# 🚛 مدیریت سفر و رهگیری محموله‌ها
# -----------------------------


# -----------------------------
# 🔔 یادآوری‌ها
# -----------------------------

@login_required
def create_reminder(request):
    if request.method == 'POST':
        form = ReminderForm(request.POST)
        if form.is_valid():
            reminder = form.save(commit=False)
            reminder.company_code = request.user.company.company_code
            reminder.save()
            messages.success(request, 'یادآوری جدید ثبت شد.')
            return redirect('reminder_list')
    else:
        form = ReminderForm()
    return render(request, 'create_reminder.html', {'form': form})


@login_required
def reminder_list(request):
    reminders = Reminder.objects.filter(company_code=request.user.company.company_code)
    return render(request, 'reminder_list.html', {'reminders': reminders})
@login_required
def subwarehouse_management(request):
    company = get_company(request.user)
    if not company:
        messages.error(request, 'لطفاً ابتدا یک شرکت ثبت کنید.')
        return redirect('register_company')
    
    subwarehouses = WarehouseSubStorage.objects.filter(company_code=company.company_code)
    ships = Ship.objects.filter(company_code=company.company_code)
    parts = ShipPart.objects.filter(ship__company_code=company.company_code)
    
    context = {
        'company': company,
        'subwarehouses': subwarehouses,
        'ships': ships,
        'parts': parts,
    }
    return render(request, 'subwarehouse_management.html', context)

@login_required
def update_subwarehouse(request, subwarehouse_id):
    try:
        company = Company.objects.get(user=request.user)
        subwarehouse = get_object_or_404(WarehouseSubStorage, id=subwarehouse_id, company_code=company.company_code)

        if request.method == 'POST':
            form = ShipSubWarehouseForm(request.POST, instance=subwarehouse)
            if form.is_valid():
                form.save()
                messages.success(request, 'مشخصات زیرانبار با موفقیت به‌روزرسانی شد.')
                return redirect('nezarat')

        else:
            form = ShipSubWarehouseForm(instance=subwarehouse)

        return render(request, 'update_subwarehouse.html', {'form': form, 'subwarehouse': subwarehouse})

    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')

@login_required
def delete_subwarehouse(request, subwarehouse_id):
    company = get_company(request.user)
    if not company:
        messages.error(request, 'لطفاً ابتدا یک شرکت ثبت کنید.')
        return redirect('register_company')
    
    subwarehouse = get_object_or_404(WarehouseSubStorage, id=subwarehouse_id, company_code=company.company_code)
    if request.method == 'POST':
        subwarehouse.delete()
        messages.success(request, 'زیرانبار حذف شد.')
        return redirect('subwarehouse_management')
    
    return render(request, 'delete_subwarehouse.html', {'subwarehouse': subwarehouse})

@login_required
def mark_worn_out(request, part_id):
    try:
        company = Company.objects.get(user=request.user)
        part = get_object_or_404(Part, id=part_id, company_code=company.company_code)

        part.is_worn_out = True
        part.save()

        messages.success(request, f'قطعه "{part.name}" به عنوان فرسوده ثبت شد.')
        return redirect('nezarat')

    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')
    
@login_required
def worn_out_parts(request):
    try:
        company = Company.objects.get(user=request.user)
        worn_parts = Part.objects.filter(company_code=company.company_code, is_worn_out=True)

        return render(request, 'worn_out_parts.html', {'worn_parts': worn_parts})

    except Company.DoesNotExist:
        messages.error(request, 'شما هنوز شرکتی ثبت نکرده‌اید.')
        return redirect('home')

from django.utils import timezone
from .models import Reminder

def reminder_notifications(request):
    if request.user.is_authenticated and hasattr(request.user, 'company'):
        now = timezone.now()
        reminders = Reminder.objects.filter(
            company_code=request.user.company.company_code,
            reminder_date__lte=now
        )
        return {'notifications': reminders}
    return {}

# views.py
@login_required
def create_transport_operation(request):
    if request.method == 'POST':
        transport_operation_form = TransportOperationForm(request.POST, user=request.user)

        TransferToShipFormSet = modelformset_factory(TransportItem, form=TransportItemForm, extra=0)
        to_ship_formset = TransferToShipFormSet(
            request.POST,
            prefix="to_ship",
            form_kwargs={'user': request.user},
            queryset=TransportItem.objects.none()  # در صورت نیاز QuerySet دلخواه بگذارید
        )

        TransferToWarehouseFormSet = modelformset_factory(TransportItem, form=TransportItemForm, extra=0)
        to_warehouse_formset = TransferToWarehouseFormSet(
            request.POST,
            prefix="to_warehouse",
            form_kwargs={'user': request.user},
            queryset=TransportItem.objects.none()
        )

        if (transport_operation_form.is_valid()
            and to_ship_formset.is_valid()
            and to_warehouse_formset.is_valid()):
            operation = transport_operation_form.save(commit=False)
            operation.company_code = request.user.company.company_code
            operation.save()

            # آیتم‌های مرحله اول
            for form in to_ship_formset:
                item = form.save(commit=False)
                item.operation = operation
                item.save()

            # آیتم‌های مرحله دوم
            for form in to_warehouse_formset:
                item = form.save(commit=False)
                item.operation = operation
                item.save()

            return redirect('transport_list')
    else:
        transport_operation_form = TransportOperationForm(user=request.user)

        TransferToShipFormSet = modelformset_factory(TransportItem, form=TransportItemForm, extra=2)
        to_ship_formset = TransferToShipFormSet(prefix="to_ship",
                                                form_kwargs={'user': request.user},
                                                queryset=TransportItem.objects.none())

        TransferToWarehouseFormSet = modelformset_factory(TransportItem, form=TransportItemForm, extra=2)
        to_warehouse_formset = TransferToWarehouseFormSet(prefix="to_warehouse",
                                                          form_kwargs={'user': request.user},
                                                          queryset=TransportItem.objects.none())

    return render(request, 'create_transport_operation.html', {
        'transport_operation_form': transport_operation_form,
        'transfer_to_ship_forms': to_ship_formset.forms,
        'transfer_to_warehouse_forms': to_warehouse_formset.forms,
    })



@login_required
def edit_transport_operation(request, operation_id):
    operation = get_object_or_404(TransportOperation, id=operation_id, company_code=request.user.company.company_code)

    if operation.status != 'PENDING':
        messages.warning(request, 'فقط عملیات در حال بررسی قابل ویرایش است.')
        return redirect('transport_detail', operation_id=operation.id)

    if request.method == 'POST':
        form = TransportOperationForm(request.POST, instance=operation, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'عملیات با موفقیت ویرایش شد.')
            return redirect('transport_detail', operation_id=operation.id)
    else:
        form = TransportOperationForm(instance=operation, user=request.user)

    return render(request, 'edit_transport_operation.html', {'form': form, 'operation': operation})

@login_required
def delete_transport_operation(request, operation_id):
    operation = get_object_or_404(TransportOperation, id=operation_id, company_code=request.user.company.company_code)

    if request.method == 'POST':
        operation.delete()
        messages.success(request, 'عملیات حمل‌ونقل حذف شد.')
        return redirect('transport_list')

    return render(request, 'delete_transport_operation.html', {'operation': operation})

@login_required
def finalize_transport_operation(request, operation_id):
    operation = get_object_or_404(TransportOperation, id=operation_id, company_code=request.user.company.company_code)

    if operation.status != 'PENDING':
        messages.warning(request, 'این عملیات دیگر در وضعیت در حال بررسی نیست.')
        return redirect('transport_detail', operation_id=operation.id)

    if request.method == 'POST':
        # شروع منطق نهایی‌کردن:
        for item in operation.items.all():
            part = item.part

            # چک کردن موجودی
            if part.quantity < item.quantity:
                messages.error(request, f"موجودی {part.name} کافی نیست.")
                return redirect('transport_detail', operation_id=operation.id)

            # کاهش موجودی از انبار مبدا
            part.quantity -= item.quantity

            # بررسی فرسودگی بر اساس مصرف
            if item.is_consumable:
                if item.usage_percent >= 100:
                    item.is_worn_out = True
                # اگر منطق دیگری دارید، اضافه کنید

            # بررسی تاریخ انقضا
            if item.will_expire_during():
                item.is_worn_out = True
                part.is_worn_out = True
                part.worn_out_reason = "انقضای قطعه در طول سفر"

            # بررسی فرسودگی با مسافت
            elif item.will_wear_out_by_distance():
                item.is_worn_out = True
                part.is_worn_out = True
                part.worn_out_reason = "فرسودگی بر اثر مسافت در سفر"

            # می‌توانید شرط‌های دیگری نیز اضافه کنید
            item.save()
            part.save()

        operation.status = 'COMPLETED'
        operation.completed_at = timezone.now()
        operation.save()

        messages.success(request, 'عملیات حمل‌ونقل نهایی شد.')
        return redirect('transport_detail', operation_id=operation.id)

    return render(request, 'finalize_transport.html', {'operation': operation})


@login_required
def add_transport_items(request, operation_id):
    operation = get_object_or_404(TransportOperation, id=operation_id, company_code=request.user.company.company_code)
    if request.method == 'POST':
        form = TransportItemForm(request.POST, user=request.user)
        if form.is_valid():
            item = form.save(commit=False)
            item.operation = operation
            item.save()
            messages.success(request, 'قطعه به عملیات اضافه شد.')
            return redirect('add_transport_items', operation_id=operation.id)
    else:
        form = TransportItemForm(user=request.user)

    items = operation.items.all()
    return render(request, 'add_transport_items.html', {'form': form, 'operation': operation, 'items': items})


@login_required
def transport_list(request):
    company = get_object_or_404(Company, user=request.user)
    query = request.GET.get('q', '')
    status = request.GET.get('status', '')

    operations = TransportOperation.objects.filter(company_code=company.company_code)

    if query:
        operations = operations.filter(ship__name__icontains=query)
    if status:
        operations = operations.filter(status=status)

    return render(request, 'transport_list.html', {
        'operations': operations,
        'query': query,
        'status': status,
    })


@login_required
def add_multiple_transport_items(request, operation_id):
    operation = get_object_or_404(TransportOperation, id=operation_id, company_code=request.user.company.company_code)

    TransportItemFormSetLocal = modelformset_factory(
        TransportItem,
        form=TransportItemForm,
        extra=3,
        can_delete=True
    )

    if request.method == 'POST':
        formset = TransportItemFormSetLocal(request.POST, queryset=TransportItem.objects.none(), form_kwargs={'user': request.user})
        if formset.is_valid():
            for form in formset:
                if form.cleaned_data and not form.cleaned_data.get('DELETE'):
                    item = form.save(commit=False)
                    item.operation = operation
                    item.save()
            messages.success(request, 'چند قطعه با موفقیت اضافه شدند.')
            return redirect('transport_detail', operation_id=operation.id)
    else:
        formset = TransportItemFormSetLocal(queryset=TransportItem.objects.none(), form_kwargs={'user': request.user})

    return render(request, 'add_multiple_transport_items.html', {'formset': formset, 'operation': operation})

@login_required
def export_transport_excel(request, operation_id):
    operation = get_object_or_404(TransportOperation, id=operation_id, company_code=request.user.company.company_code)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="transport_{operation.id}.csv"'
    writer = csv.writer(response)
    
    writer.writerow(['قطعه', 'تعداد', 'زیرانبار کشتی', 'مصرفی', 'استفاده شده', 'فرسوده'])

    for item in operation.items.all():
        writer.writerow([
            item.part.name,
            item.quantity,
            item.ship_subwarehouse.name if item.ship_subwarehouse else "-",
            'بله' if item.is_consumable else 'خیر',
            '✅' if item.used_in_trip else '❌',
            '🔴' if item.is_worn_out else '🟢',
        ])
    
    return response

from django.template.loader import render_to_string

@login_required
def export_transport_pdf(request, operation_id):
    operation = get_object_or_404(TransportOperation, id=operation_id, company_code=request.user.company.company_code)
    template = 'pdf/transport_report.html'
    html = render_to_string(template, {'operation': operation})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="transport_{operation.id}.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response

@login_required
def approve_transport_operation(request, operation_id):
    operation = get_object_or_404(TransportOperation, id=operation_id, ship__captain=request.user)

    if request.method == 'POST':
        operation.is_approved_by_captain = True
        operation.approved_at = timezone.now()
        operation.status = 'IN_PROGRESS'
        operation.save()
        messages.success(request, 'سفر تأیید شد و در حال حرکت است.')
        return redirect('transport_detail', operation_id=operation.id)

    return render(request, 'approve_transport.html', {'operation': operation})

@login_required
def transport_detail(request, operation_id):
    operation = get_object_or_404(TransportOperation, id=operation_id, company_code=request.user.company.company_code)
    items = operation.items.all()
    expired_items = [item for item in items if item.will_expire_during()]
    return render(request, 'transport_detail.html', {
        'operation': operation,
        'items': items,
        'expired_items': expired_items,
    })

@login_required
def create_warehouse_substorage(request):
    try:
        company = Company.objects.get(user=request.user)
        if request.method == 'POST':
            form = WarehouseSubStorageForm(request.POST, user=request.user)
            if form.is_valid():
                form.save()
                messages.success(request, 'زیرانبار جدید برای انبار با موفقیت ثبت شد.')
                return redirect('warehouse_list')
        else:
            form = WarehouseSubStorageForm(user=request.user)
        return render(request, 'create_warehouse_substorage.html', {'form': form})
    except Company.DoesNotExist:
        messages.error(request, 'ابتدا باید شرکت ثبت کنید.')
        return redirect('home')

@login_required
def create_ship_subwarehouse(request):
    try:
        company = Company.objects.get(user=request.user)
        if request.method == 'POST':
            form = ShipSubWarehouseForm(request.POST, user=request.user)
            if form.is_valid():
                form.save()
                messages.success(request, 'زیرانبار جدید برای کشتی با موفقیت ثبت شد.')
                return redirect('company_dashboard')
        else:
            form = ShipSubWarehouseForm(user=request.user)
        return render(request, 'create_ship_subwarehouse.html', {'form': form})
    except Company.DoesNotExist:
        messages.error(request, 'ابتدا شرکت خود را ثبت کنید.')
        return redirect('home')

@login_required
def add_ship_attribute(request):
    try:
        company = Company.objects.get(user=request.user)
        if request.method == 'POST':
            form = ShipAttributeForm(request.POST, user=request.user)
            if form.is_valid():
                form.save()
                messages.success(request, 'ویژگی جدید برای کشتی ثبت شد.')
                return redirect('company_dashboard')
        else:
            form = ShipAttributeForm(user=request.user)
        return render(request, 'add_ship_attribute.html', {'form': form})
    except Company.DoesNotExist:
        messages.error(request, 'ابتدا باید شرکت ثبت کنید.')
        return redirect('home')

def unread_message_count(request):
    if request.user.is_authenticated:
        return {
            'unread_count': Message.objects.filter(receiver=request.user, is_read=False).count()
        }
    return {}

@login_required
def report_damage(request):
    if request.method == 'POST':
        form = DamageReportForm(request.POST, user=request.user)
        if form.is_valid():
            report = form.save(commit=False)
            report.reported_by = request.user
            report.save()
            messages.success(request, 'گزارش خرابی ارسال شد و در انتظار تایید مدیر است.')
            return redirect('home')
    else:
        form = DamageReportForm(user=request.user)

    return render(request, 'report_damage.html', {'form': form})

@login_required
def review_damage_reports(request):
    if not request.user.title == 'C':
        return HttpResponseForbidden()

    reports = DamageReport.objects.filter(status='PENDING')
    return render(request, 'review_damage_reports.html', {'reports': reports})

@login_required
def confirm_damage_report(request, report_id):
    report = get_object_or_404(DamageReport, id=report_id)
    if request.user.title != 'C':
        return HttpResponseForbidden()

    part = report.part
    if part.quantity > 0:
        part.quantity -= 1
        part.is_worn_out = True
        part.worn_out_reason = "توسط کاربر گزارش شده"
        part.save()
        report.status = 'CONFIRMED'
        report.save()
        messages.success(request, 'گزارش تایید شد و قطعه فرسوده شد.')
    else:
        messages.warning(request, 'قطعه‌ای برای فرسودگی باقی نمانده.')

    return redirect('review_damage_reports')

@login_required
def part_status_report(request):
    company = get_object_or_404(Company, user=request.user)
    parts = Part.objects.filter(company_code=company.company_code)

    part_statuses = []
    for part in parts:
        part_statuses.append({
            'part': part,
            'status': part.get_status()
        })

    return render(request, 'part_status_report.html', {'part_statuses': part_statuses})

@login_required
def role_based_dashboard(request):
    user = request.user

    if user.title == 'C':  # کاپیتان
        employee = get_object_or_404(Employee, user=user)
        ship = employee.assigned_ship
        return render(request, 'captain_dashboard.html', {
            'ship': ship,
            'subwarehouses': ship.sub_warehouses.all(),
            'parts': ShipPart.objects.filter(ship=ship),
        })

    elif user.title == 'OS':  # اپراتور کشتی
        employee = get_object_or_404(Employee, user=user)
        ship = employee.assigned_ship
        return render(request, 'os_dashboard.html', {
            'ship': ship,
            'parts': ShipPart.objects.filter(ship=ship),
        })

    elif user.title in ['OA', 'OB']:  # اپراتور انبار
        employee = get_object_or_404(Employee, user=user)
        warehouse = employee.assigned_warehouse
        return render(request, 'warehouse_operator_dashboard.html', {
            'warehouse': warehouse,
            'substorages': warehouse.sub_storages.all(),
            'parts': Part.objects.filter(warehouse=warehouse),
        })

    elif user.title == 'E':  # کارمند عادی
        return render(request, 'basic_user_dashboard.html')

    else:
        return redirect('home')
@login_required
def review_damage_reports(request):
    if request.user.title != 'C':
        return HttpResponseForbidden()

    reports = DamageReport.objects.filter(status='PENDING')
    return render(request, 'review_damage_reports.html', {'reports': reports})
@login_required
def handle_damage_report(request, report_id, action):
    report = get_object_or_404(DamageReport, id=report_id)
    part = report.part

    if request.user.title != 'C':
        return HttpResponseForbidden()

    if action == 'confirm':
        if part.quantity > 0:
            part.quantity -= 1
            part.is_worn_out = True
            part.worn_out_reason = "توسط مدیر تایید شده"
            part.save()
            report.status = 'CONFIRMED'
            report.save()
            messages.success(request, f'قطعه {part.name} فرسوده اعلام شد.')
    elif action == 'reject':
        report.status = 'REJECTED'
        report.save()
        messages.info(request, 'درخواست رد شد.')

    return redirect('review_damage_reports')

@login_required
def ship_visual_dashboard(request, ship_id):
    ship = get_object_or_404(Ship, id=ship_id)
    subwarehouses = ShipSubWarehouse.objects.filter(ship=ship)
    return render(request, 'ship_visual_dashboard.html', {'ship': ship, 'subwarehouses': subwarehouses})

@login_required
def chat_list(request):
    users = User.objects.exclude(id=request.user.id)
    return render(request, 'chat_list.html', {'users': users})

def create_part(request):
    if request.method == 'POST':
        form = PartForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()  # ذخیره قطعه جدید
            return redirect('part_list')  # بعد از ذخیره به صفحه لیست قطعات هدایت می‌شود
    else:
        form = PartForm()

    return render(request, 'create_part.html', {'form': form})


def transport_operation_list(request):
    operations = TransportOperation.objects.all()  # لیست تمام عملیات حمل و نقل
    return render(request, 'transport_operation_list.html', {'operations': operations})

def part_list(request):
    parts = Part.objects.all()  # لیست تمام قطعات
    return render(request, 'part_list.html', {'parts': parts})

def create_warehouse_with_substorages(request):
    """
    ساخت یک انبار به همراه هر تعداد زیرانبار.
    """
    if request.method == 'POST':
        warehouse_form = WarehouseForm(request.POST)
        formset = WarehouseSubStorageFormSet(request.POST)
        if warehouse_form.is_valid() and formset.is_valid():
            # مرحله ۱) انبار را ذخیره کن
            warehouse = warehouse_form.save()

            # مرحله ۲) زیرانبارها را ذخیره کن
            # commit=False یعنی هنوز ذخیره نشود تا بتوانیم
            # فیلد warehouseشان را ست کنیم
            subs = formset.save(commit=False)
            for sub in subs:
                sub.warehouse = warehouse
                sub.save()

            # اگر کاربر بعضی فرم‌ها را تیک حذف زده باشد:
            for del_form in formset.deleted_forms:
                if del_form.instance.pk:
                    del_form.instance.delete()

            messages.success(request, "انبار جدید و زیرانبارها با موفقیت ذخیره شدند.")
            return redirect('warehouse_list')
    else:
        warehouse_form = WarehouseForm()
        formset = WarehouseSubStorageFormSet()

    return render(request, 'create_warehouse_with_substorages.html', {
        'warehouse_form': warehouse_form,
        'formset': formset
    })
    
def ship_details(request, ship_id):
    """جزئیات یک کشتی + زیرانبارهای کشتی + قطعات داخل هر زیرانبار کشتی"""
    ship = get_object_or_404(Ship, id=ship_id)

    # بررسی کنید کشتی مربوط به شرکت لاگین‌کرده باشد
    try:
        company = Company.objects.get(user=request.user)
        if ship.company_code != company.company_code:
            return HttpResponseForbidden("شما اجازه دسترسی به این کشتی را ندارید.")
    except Company.DoesNotExist:
        return HttpResponseForbidden("شما شرکت ثبت نکرده‌اید.")

    # زیرانبارهای مربوط به این کشتی
    subwarehouses = ShipSubWarehouse.objects.filter(ship=ship)
    # قطعات هر زیرانبار
    subwarehouse_parts = {
        sw: Part.objects.filter(ship_subwarehouse=sw)
        for sw in subwarehouses
    }

    context = {
        'ship': ship,
        'subwarehouses': subwarehouses,
        'subwarehouse_parts': subwarehouse_parts,
    }
    return render(request, 'ship_detail.html', context)


@login_required
def edit_ship_modal(request, ship_id):
    """
    نمونه ویوی Ajax برای ویرایش کشتی (Modal Edit).
    """
    ship = get_object_or_404(Ship, id=ship_id)
    # محدودیت شرکت
    try:
        company = Company.objects.get(user=request.user)
        if ship.company_code != company.company_code:
            return HttpResponseForbidden("اجازه دسترسی ندارید.")
    except Company.DoesNotExist:
        return HttpResponseForbidden("شرکت شما معتبر نیست.")

    if request.method == 'POST':
        form = ShipForm(request.POST, request.FILES, instance=ship)
        if form.is_valid():
            form.save()
            return JsonResponse({'status': 'ok', 'message': 'کشتی با موفقیت ویرایش شد.'})
        else:
            # فرم با خطاها برگردد
            return render(request, 'partials/ship_edit_form.html', {'form': form})
    else:
        form = ShipForm(instance=ship)
        return render(request, 'partials/ship_edit_form.html', {'form': form})


@login_required
def delete_ship_ajax(request, ship_id):
    """
    نمونه ویوی Ajax برای حذف کشتی
    """
    if request.method == 'POST':
        try:
            company = Company.objects.get(user=request.user)
            ship = Ship.objects.get(id=ship_id, company_code=company.company_code)
            ship.delete()
            return JsonResponse({'status': 'ok'})
        except:
            return JsonResponse({'status': 'error'}, status=400)
    return JsonResponse({'status': 'invalid'}, status=405)
